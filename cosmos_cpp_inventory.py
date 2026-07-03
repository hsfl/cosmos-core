#!/usr/bin/env python3
"""Generate COSMOS core .cpp file inventory spreadsheet.
   Reads user-filled Category/Sub-Category from existing Programs tab before overwriting.

   Layer scheme: Kernel (K0–K5) and Support (S0–S4) each number from 0.
   Fresh deps from analyze_deps.py run 2026-06-17.
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = "/home2/pilger/cosmos/src/core"
EXISTING = os.path.join(BASE, "cosmos_cpp_inventory.xlsx")

# ─── Read user data from existing spreadsheet ────────────────────────────────
user_program_data = {}
if os.path.exists(EXISTING):
    wb_old = openpyxl.load_workbook(EXISTING)
    if "Programs" in wb_old.sheetnames:
        ws_old = wb_old["Programs"]
        hdrs = [c.value for c in ws_old[1]]
        ci = {h: i for i, h in enumerate(hdrs) if h}
        for row in ws_old.iter_rows(min_row=2, values_only=True):
            fname = row[0]
            if not fname or str(fname).startswith("LEGEND"):
                continue
            cat     = row[ci.get("Category", 2)]     if "Category" in ci else None
            sub_cat = row[ci.get("Sub-Category", 3)] if "Sub-Category" in ci else None
            user_program_data[fname] = (cat, sub_cat)

# ─── Colour palette ──────────────────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
HDR_FONT  = Font(bold=True, color="FFFFFF", size=11)
WARN_FILL = PatternFill("solid", fgColor="FFE699")   # amber – open issue

# Kernel layers K0–K5: light → dark green
KERNEL_COLORS = {
    "K0": "F0F7EE", "K1": "D6EAD0", "K1": "BBE0B0",
    "K2": "9DD490", "K3": "7EC873", "K4": "5CB85C",
}
# Support layers S0–S4: light → dark orange
SUPPORT_COLORS = {
    "S0": "FEF0E7", "S1": "FDDBB8", "S2": "FCC48A",
    "S3": "FBAC5C", "S4": "F4912E",
}

thin   = Side(style="thin", color="AAAAAA")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def layer_fill(layer_str):
    color = KERNEL_COLORS.get(layer_str) or SUPPORT_COLORS.get(layer_str, "FFFFFF")
    return PatternFill("solid", fgColor=color)

# ─── Library data ─────────────────────────────────────────────────────────────
# (filename_no_ext, rel_dir, group, sub_group, layer, kernel_support, key_deps, flag)
# layer:          "K0"–"K4" (kernel) | "S0"–"S4" (support)
# kernel_support: "kernel" | "support"
# flag:           "open" (known issue still open) | ""
LIBRARIES = [
    # ── MATH  K0 ─────────────────────────────────────────────────────────────
    ("bindings",         "libraries/math",    "math","—",        "K0","kernel","vector.h",""),
    ("bytelib",          "libraries/math",    "math","—",        "K0","kernel","—",""),
    ("crclib",           "libraries/math",    "math","—",        "K0","kernel","—",""),
    ("lsfit",            "libraries/math",    "math","—",        "K0","kernel","—",""),
    ("mathlib",          "libraries/math",    "math","—",        "K0","kernel","—",""),
    ("matrix",           "libraries/math",    "math","—",        "K0","kernel","—",""),
    ("rotation",         "libraries/math",    "math","—",        "K0","kernel","—",""),
    ("vector",           "libraries/math",    "math","—",        "K0","kernel","matrix.h",""),

    # ── SUPPORT  K0 – primitives ──────────────────────────────────────────────
    ("cosmos-errclass",  "libraries/support","support","errors",   "K0","kernel","cosmos-errclass.h (timelib.h ⚠ possibly unused)",""),
    ("json11",           "libraries/support","support","json",     "K0","kernel","—",""),
    ("jsonclass",        "libraries/support","support","json",     "K0","kernel","—",""),
    ("jsonobject",       "libraries/support","support","json",     "K0","kernel","—",""),
    ("jsonvalue",        "libraries/support","support","json",     "K0","kernel","jsonobject.h, stringlib.h",""),
    ("print_utils",      "libraries/support","support","utilities","K0","kernel","configCosmos.h",""),
    ("sliplib",          "libraries/support","support","network",  "K0","kernel","—",""),

    # ── SUPPORT  K1 – core utilities ─────────────────────────────────────────
    ("elapsedtime",      "libraries/support","support","time",     "K0","kernel","timelib.h (currentmjd, secondsleep)",""),
    ("stringlib",        "libraries/support","support","utilities","K0","kernel","mathlib.h, jsonobject.h",""),
    ("timelib",          "libraries/support","support","time",     "K0","kernel","timeutils.h, stringlib.h, mathlib.h (⚠ possibly unused)",""),
    ("timeutils",        "libraries/support","support","time",     "K0","kernel","—",""),

    # ── SUPPORT  K2 – data I/O ───────────────────────────────────────────────
    ("ax25class",        "libraries/support","support","network",  "K1","kernel","configCosmosKernel.h",""),
    ("check",            "libraries/support","support","utilities","K1","kernel","timelib.h",""),
    ("datalib",          "libraries/support","support","data",     "K1","kernel","jsondef.h, stringlib.h, timelib.h",""),
    ("logger",           "libraries/support","support","utilities","K1","kernel","data_base_path (output only)",""),
    ("socketlib",        "libraries/support","support","network",  "K1","kernel","elapsedtime.h, timelib.h, bytelib.h",""),

    # ── SUPPORT  K3 – JSON / namespace ───────────────────────────────────────
    ("jsondef",          "libraries/support","support","json",     "K2","kernel","cosmos-errno.h",""),
    ("objlib",           "libraries/support","support","data",     "K2","kernel","datalib.h",""),
    ("constants",        "libraries/physics", "physics","—",       "K2","kernel","—",""),

    # ── SUPPORT  K4 – domain / orbit foundations ─────────────────────────────
    ("channellib",       "libraries/support","support","network",  "K3","kernel","timelib.h",""),
    ("transferclass",    "libraries/support","support","network",  "K3","kernel","data_base_path (output only)",""),
    ("transferlib",      "libraries/support","support","network",  "K3","kernel","bytelib.h",""),

    # ── SUPPORT  K5 – messaging & devices ────────────────────────────────────
    ("packetcomm",       "libraries/support","support","messaging","K4","kernel","ax25class.h, crclib.h, sliplib.h, timelib.h",""),
    ("FileSender",       "libraries/support/filesenderimpl","support","messaging","K4","kernel","transferclass.h",""),
    ("UdpSender",        "libraries/support/filesenderimpl","support","messaging","K4","kernel","socketlib.h, transferclass.h",""),
    ("task",             "libraries/agent",  "agent","utilities",  "K4","kernel","datalib.h, timelib.h",""),
    ("arduino_lib",      "libraries/device/arduino",  "device","arduino",    "K4","kernel","configCosmos.h, timelib.h",""),
    ("spp",              "libraries/device/ccsds",    "device","ccsds",      "K4","kernel","—",""),
    ("devicecpu",        "libraries/device/cpu",      "device","cpu",        "K4","kernel","—",""),
    ("devicedisk",       "libraries/device/disk",     "device","disk",       "K4","kernel","—",""),
    ("bbFctns",          "libraries/device/general",  "device","beaglebone", "K4","kernel","cssl_lib.h",""),
    ("cssl_lib",         "libraries/device/general",  "device","serial",     "K4","kernel","configCosmos.h",""),
    ("gige_lib",         "libraries/device/general",  "device","camera",     "K4","kernel","mathlib.h, timelib.h",""),
    ("gs232b_lib",       "libraries/device/general",  "device","antenna",    "K4","kernel","configCosmos.h",""),
    ("ic9100_lib",       "libraries/device/general",  "device","radio",      "K4","kernel","jsondef.h",""),
    ("kisslib",          "libraries/device/general",  "device","TNC",        "K4","kernel","configCosmos.h",""),
    ("kisstnc_lib",      "libraries/device/general",  "device","TNC",        "K4","kernel","—",""),
    ("kpc9612p_lib",     "libraries/device/general",  "device","TNC",        "K4","kernel","timelib.h",""),
    ("mixwtnc_lib",      "libraries/device/general",  "device","TNC",        "K4","kernel","—",""),
    ("pic_lib",          "libraries/device/general",  "device","PIC MCU",    "K4","kernel","crclib.h",""),
    ("prkx2su_class",    "libraries/device/general",  "device","antenna",    "K4","kernel","—",""),
    ("prkx2su_lib",      "libraries/device/general",  "device","antenna",    "K4","kernel","configCosmos.h",""),
    ("ts2000_lib",       "libraries/device/general",  "device","radio",      "K4","kernel","configCosmos.h",""),
    ("unixgpio",         "libraries/device/general",  "device","GPIO",       "K4","kernel","datalib.h",""),
    ("usrp_lib",         "libraries/device/general",  "device","SDR",        "K4","kernel","jsondef.h",""),
    ("i2c",              "libraries/device/i2c",      "device","I2C",        "K4","kernel","—",""),
    ("netradio",         "libraries/device/netradio", "device","radio",      "K4","kernel","timelib.h",""),
    ("serialclass",      "libraries/device/serial",   "device","serial",     "K4","kernel","configCosmos.h",""),

    # ── PHYSICS / SUPPORT  S0 – high-level foundations ───────────────────────
    ("convertlib",       "libraries/support","support","coordinate","S0","support","ephemlib.h, geomag.h → get_cosmosresources()",""),
    ("demlib",           "libraries/support","support","data",      "S0","support","get_cosmosresources() – DEM tiles",""),
    ("enumlib",          "libraries/support","support","utilities", "S0","support","convertlib.h → get_cosmosresources()",""),
    ("envi",             "libraries/support","support","data",      "S0","support","datalib.h, bytelib.h (reads/writes image files)",""),
    ("ephemlib",         "libraries/support","support","ephemeris", "S0","support","get_cosmosresources() – JPL ephemeris",""),
    ("geomag",           "libraries/support","support","ephemeris", "S0","support","get_cosmosresources() – WMM.COF",""),
    ("jpleph",           "libraries/support","support","ephemeris", "S0","support","fopen/fread – JPL binary ephemeris file",""),
    ("jsonlib",          "libraries/support","support","json",      "S0","support","jsonlib.h, convertlib.h (cosmos-errno.h ⚠ possibly unused)",""),
    ("nrlmsise-00",      "libraries/physics","physics","atmosphere","S0","support","physics simulation context only",""),
    ("nrlmsise-00_data", "libraries/physics","physics","atmosphere","S0","support","—",""),
    ("physicslib",       "libraries/physics","physics","orbital",   "S0","support","get_cosmosresources() – resource files",""),
    # convert_test_gui (Qt GUI test app – should move to programs/tests/)
    ("eci2kep_test",     "libraries/support/convert_test_gui","support","GUI test","S0","support","convertlib.h",""),
    ("filepathdialog",   "libraries/support/convert_test_gui","support","GUI test","S0","support","—",""),
    ("main (gui)",       "libraries/support/convert_test_gui","support","GUI test","S0","support","—",""),
    ("mainwindow",       "libraries/support/convert_test_gui","support","GUI test","S0","support","—",""),
    ("rearth_test",      "libraries/support/convert_test_gui","support","GUI test","S0","support","—",""),
    ("testcontainer",    "libraries/support/convert_test_gui","support","GUI test","S0","support","—",""),

    # ── PHYSICS  S1 – physics simulation ─────────────────────────────────────
    ("controllib",       "libraries/physics","physics","control",   "S1","support","physicsclass.h → get_cosmosresources()",""),
    ("physicsclass",     "libraries/physics","physics","simulation","S1","support","get_cosmosresources() – spacecraft model files",""),
    ("simulatorclass",   "libraries/physics","physics","simulation","S1","support","physicsclass.h → get_cosmosresources()",""),

    # ── AGENT  S2 – agent framework ──────────────────────────────────────────
    ("agentclass",       "libraries/agent","agent","—",            "S2","support","jsonlib.h → convertlib → get_cosmosresources()",""),

    # ── AGENT / SUPPORT  S3 – agent utilities & devices ─────────────────────
    # beacon/packethandler: agentclass removed from beacon.h by user cleanup,
    # but beacon still includes jsonlib.h and enumlib.h (both S0) → S3
    # packethandler still includes agentclass.h and convertlib.h → S3
    ("beacon",           "libraries/support","support","messaging", "S3","support","jsonlib.h, enumlib.h, packetcomm.h",""),
    ("packethandler",    "libraries/support","support","messaging", "S3","support","agentclass.h, convertlib.h, beacon.h, packetcomm.h",""),
    ("command_queue",    "libraries/agent","agent","utilities",    "S3","support","agentclass.h, event.h, task.h",""),
    ("event",            "libraries/agent","agent","utilities",    "S3","support","jsonlib.h (⚠ possibly unused), timelib.h",""),
    ("scheduler",        "libraries/agent","agent","utilities",    "S3","support","agentclass.h, event.h",""),
    ("acq_a35",          "libraries/device/general","device","camera","S3","support","agentclass.h → get_cosmosresources()",""),

    # ── MODULE  S4 ────────────────────────────────────────────────────────────
    ("file_module",            "libraries/module","module","—","S4","support","agentclass → get_cosmosresources()",""),
    ("node_propagator_module", "libraries/module","module","—","S4","support","physicsclass → get_cosmosresources()",""),
    ("packethandler_module",   "libraries/module","module","—","S4","support","agentclass → get_cosmosresources()",""),
    ("websocket_module",       "libraries/module","module","—","S4","support","agentclass → get_cosmosresources()",""),
]

# ─── Program data ─────────────────────────────────────────────────────────────
PROGRAMS = [
    ("agent",              "programs/agents",               "support","agentclass, physicslib, datalib, jsonlib"),
    ("agent_cpu",          "programs/agents",               "support","agentclass, devicecpu, devicedisk, print_utils"),
    ("agent_data",         "programs/agents",               "support","agentclass, physicslib, datalib, jsonlib"),
    ("agent_exec",         "programs/agents",               "support","agentclass, command_queue, convertlib, datalib"),
    ("agent_file",         "programs/agents",               "support","agentclass, file_module, websocket_module, packethandler"),
    ("agent_forward",      "programs/agents",               "support","agentclass"),
    ("agent_monitor",      "programs/agents",               "support","agentclass, command_queue, convertlib, jsonlib"),
    ("agent_propagator",   "programs/agents",               "support","agentclass, simulatorclass, beacon, packetcomm, packethandler"),
    ("agent_route",        "programs/agents",               "support","agentclass"),
    ("agent_time",         "programs/agents",               "support","agentclass, command_queue, convertlib, jsonlib"),
    ("agent_tunnel",       "programs/agents",               "support","agentclass, serialclass"),
    ("agent_tunnel2",      "programs/agents",               "support","agentclass, serialclass"),
    ("add_radio",          "programs/agents/ground-station","support","agentclass, ic9100_lib"),
    ("agent_antenna",      "programs/agents/ground-station","support","agentclass, gs232b_lib, prkx2su_lib, convertlib"),
    ("agent_control",      "programs/agents/ground-station","support","agentclass, physicslib, ephemlib, jsonlib"),
    ("agent_kpc9612p",     "programs/agents/ground-station","support","agentclass, cssl_lib, kpc9612p_lib, socketlib"),
    ("agent_radio",        "programs/agents/ground-station","support","agentclass, ic9100_lib, ts2000_lib, usrp_lib"),
    ("ax25_recv",          "programs/agents/ground-station","kernel", "ax25class, serialclass, socketlib"),
    ("ic9100",             "programs/agents/ground-station","support","agentclass, ic9100_lib, usrp_lib"),
    ("kiss_recv",          "programs/agents/ground-station","kernel", "kisslib, serialclass, socketlib"),
    ("kiss_send",          "programs/agents/ground-station","kernel", "kisslib, serialclass, timelib"),
    ("kpc9612p_recv",      "programs/agents/ground-station","kernel", "kpc9612p_lib, timelib"),
    ("kpc9612p_send",      "programs/agents/ground-station","kernel", "kpc9612p_lib, timelib"),
    ("monitor_antenna",    "programs/agents/ground-station","support","agentclass"),
    ("monitor_gs",         "programs/agents/ground-station","support","agentclass"),
    ("agent_arduino",      "programs/agents/other",         "support","agentclass, devicecpu, cssl_lib"),
    ("agent_node",         "programs/agents/other",         "support","agentclass, physicslib, jsonlib"),
    ("agent_physics",      "programs/agents/other",         "support","agentclass, mathlib, physicslib, socketlib"),
    ("agent_transmitter",  "programs/agents/other",         "support","agentclass, cssl_lib, kisslib, physicslib"),
    ("agent_transmitter2", "programs/agents/other",         "support","agentclass, cssl_lib, kisslib, physicslib"),
    ("archive",            "programs/general",              "support","agentclass, datalib, timelib"),
    ("calc_transform",     "programs/general",              "kernel", "mathlib"),
    ("check_satnogs",      "programs/general",              "support","agentclass, convertlib, jsonclass"),
    ("command_generator",  "programs/general",              "support","event, scheduler, jsonlib, timelib"),
    ("command_generator_remote","programs/general",         "support","event, scheduler, jsonlib, timelib"),
    ("cosmos_size",        "programs/general",              "support","agentclass"),
    ("countdown",          "programs/general",              "kernel", "timelib, elapsedtime"),
    ("cubesat2obj",        "programs/general",              "kernel", "objlib"),
    ("devstruc_size",      "programs/general",              "kernel", "jsondef"),
    ("eci2lvlh",           "programs/general",              "support","agentclass, physicslib, jsonlib"),
    ("fast_contacts",      "programs/general",              "support","agentclass, physicslib, convertlib, mathlib"),
    ("fast_propagator",    "programs/general",              "support","agentclass, physicslib, convertlib, datalib"),
    ("geoc2tle",           "programs/general",              "support","agentclass, physicslib, jsonlib"),
    ("get_contacts",       "programs/general",              "support","agentclass, physicslib, convertlib"),
    ("get_contacts_tle",   "programs/general",              "support","agentclass, physicslib, convertlib"),
    ("get_ground_contacts","programs/general",              "support","agentclass, physicslib, convertlib"),
    ("gige_ffc",           "programs/general",              "kernel", "gige_lib"),
    ("gige_list",          "programs/general",              "support","agentclass, gige_lib, datalib"),
    ("gige_snap",          "programs/general",              "support","agentclass, gige_lib, envi"),
    ("i2ctalk",            "programs/general",              "kernel", "i2c, datalib, logger"),
    ("initialize_time",    "programs/general",              "support","agentclass, timelib"),
    ("json2tab",           "programs/general",              "kernel", "jsonclass, jsonobject"),
    ("julian",             "programs/general",              "support","convertlib, timelib"),
    ("latest_file",        "programs/general",              "support","agentclass, datalib, timelib"),
    ("list_namespace",     "programs/general",              "support","agentclass, jsondef"),
    ("make_path",          "programs/general",              "kernel", "datalib, timelib"),
    ("mjd",                "programs/general",              "kernel", "timelib"),
    ("netperf_listen",     "programs/general",              "support","agentclass, sliplib, timelib"),
    ("netperf_send",       "programs/general",              "support","agentclass, sliplib, timelib"),
    ("propagatorv2",       "programs/general",              "support","agentclass, simulatorclass, jsonclass"),
    ("propagatorv3",       "programs/general",              "support","agentclass, simulatorclass, devicecpu, devicedisk"),
    ("propagatorv4",       "programs/general",              "support","agentclass, simulatorclass, devicecpu, devicedisk"),
    ("propagatorvx",       "programs/general",              "support","agentclass, simulatorclass"),
    ("propagator_web_json","programs/general",              "support","simulatorclass, cosmos-errclass, datalib"),
    ("serial_listen",      "programs/general",              "kernel", "serialclass, logger, elapsedtime"),
    ("serial_talk",        "programs/general",              "kernel", "serialclass, timelib"),
    ("shift_solid_thermal","programs/general",              "support","convertlib, jsonclass, timelib"),
    ("state2tle",          "programs/general",              "support","agentclass, physicslib, simulatorclass"),
    ("tab2json",           "programs/general",              "kernel", "jsonclass, jsonobject, stringlib"),
    ("targetsim",          "programs/general",              "support","agentclass, simulatorclass, jsonclass"),
    ("tle2state",          "programs/general",              "support","convertlib, jsonlib"),
    ("udp_listen",         "programs/general",              "support","agentclass, convertlib, jsonlib"),
    ("udp_request",        "programs/general",              "support","agentclass, jsondef, sliplib"),
    ("udp_send",           "programs/general",              "support","agentclass, jsondef, sliplib"),
    ("track_sband",        "programs/ground-station",       "support","agentclass, gs232b_lib, prkx2su_lib, convertlib"),
    ("agent_simple_request","programs/tests",               "support","agentclass, elapsedtime, timeutils"),
    ("attlvlh",            "programs/tests",                "support","physicslib, convertlib"),
    ("check_check",        "programs/tests",                "kernel", "check, jsonclass"),
    ("crctest",            "programs/tests",                "kernel", "crclib, stringlib"),
    ("errortest",          "programs/tests",                "kernel", "logger, timelib"),
    ("filespeed",          "programs/tests",                "kernel", "—"),
    ("gauss_jackson_test", "programs/tests",                "support","agentclass, simulatorclass"),
    ("geod2eci2geod",      "programs/tests",                "support","convertlib, timelib"),
    ("file_it",            "programs/tests/integration_tests/file","kernel","datalib"),
    ("file_it_helpers",    "programs/tests/integration_tests/file","kernel","datalib"),
    ("integration_tests",  "programs/tests/integration_tests",    "kernel","—"),
    ("log_move_test",      "programs/tests",                "kernel", "datalib, elapsedtime"),
    ("lsfittest",          "programs/tests",                "kernel", "mathlib"),
    ("mathspeed",          "programs/tests",                "kernel", "devicecpu, elapsedtime"),
    ("netspeed",           "programs/tests",                "kernel", "—"),
    ("netspeedd",          "programs/tests",                "kernel", "—"),
    ("objread",            "programs/tests",                "kernel", "objlib"),
    ("observation_windows","programs/tests",                "support","convertlib, jsondef, stringlib"),
    ("poslvlh",            "programs/tests",                "support","physicslib, convertlib"),
    ("serialPutData",      "programs/tests",                "kernel", "serialclass"),
    ("serialSendChar",     "programs/tests",                "kernel", "serialclass"),
    ("serial_setdtr",      "programs/tests",                "kernel", "serialclass, timelib"),
    ("string_float_test",  "programs/tests",                "support","convertlib, jsondef, json11, stringlib"),
    ("string_test",        "programs/tests",                "kernel", "stringlib"),
    ("targetstruc_tests",  "programs/tests",                "support","agentclass, simulatorclass, physicsclass, convertlib"),
    ("test_coordinate_transformations","programs/tests",    "support","convertlib, jsondef, timelib"),
    ("testdata",           "programs/tests",                "kernel", "—"),
    ("test_device_i2c",    "programs/tests",                "kernel", "i2c"),
    ("test_event_scheduler","programs/tests",               "support","convertlib, jsondef, timelib"),
    ("testmath1",          "programs/tests",                "kernel", "mathlib, vector"),
    ("tle2orbit",          "programs/tests",                "support","convertlib, jsonlib"),
    ("tledump",            "programs/tests",                "support","convertlib"),
    ("tletest",            "programs/tests",                "support","convertlib"),
    ("track_sband (test)", "programs/tests",                "support","agentclass, gs232b_lib, prkx2su_class, physicslib, convertlib"),
    ("agent_ut",           "programs/tests/unit_tests/libraries/agent",             "support","agentclass"),
    ("beacon_ut",          "programs/tests/unit_tests/libraries/support/beacon",    "support","beacon"),
    ("channellib_ut",      "programs/tests/unit_tests/libraries/support/channel",   "kernel", "channellib"),
    ("convertlib_ut",      "programs/tests/unit_tests/libraries/support/convert",   "support","convertlib"),
    ("ephemlib_ut",        "programs/tests/unit_tests/libraries/support/convert",   "support","ephemlib"),
    ("jpleph_ut",          "programs/tests/unit_tests/libraries/support/convert",   "support","jpleph"),
    ("datalib_ut",         "programs/tests/unit_tests/libraries/support/data",      "kernel", "datalib"),
    ("jsonclass_ut",       "programs/tests/unit_tests/libraries/support/json",      "kernel", "jsonclass"),
    ("logger_ut",          "programs/tests/unit_tests/libraries/support/log",       "kernel", "logger"),
    ("jsondef_ut",         "programs/tests/unit_tests/libraries/support/namespace", "kernel", "jsondef"),
    ("jsonlib_ut",         "programs/tests/unit_tests/libraries/support/namespace", "support","jsonlib"),
    ("socketlib_ut",       "programs/tests/unit_tests/libraries/support/network",   "kernel", "socketlib"),
    ("packethandler_ut",   "programs/tests/unit_tests/libraries/support/packethandler","support","packethandler"),
    ("packetcomm_ut",      "programs/tests/unit_tests/libraries/support/packet",    "kernel", "packetcomm"),
    ("print_utils_ut",     "programs/tests/unit_tests/libraries/support/print",     "kernel", "print_utils"),
    ("sliplib_ut",         "programs/tests/unit_tests/libraries/support/slip",      "kernel", "sliplib"),
    ("stringlib_ut",       "programs/tests/unit_tests/libraries/support/string",    "kernel", "stringlib"),
    ("elapsedtime_ut",     "programs/tests/unit_tests/libraries/support/time",      "kernel", "elapsedtime"),
    ("timelib_ut",         "programs/tests/unit_tests/libraries/support/time",      "kernel", "timelib"),
    ("transferclass_ut",   "programs/tests/unit_tests/libraries/support/transfer",  "kernel", "transferclass"),
    ("transferlib_ut",     "programs/tests/unit_tests/libraries/support/transfer",  "kernel", "transferlib"),
]

# ─── Helpers ──────────────────────────────────────────────────────────────────
def set_header_row(ws, headers, row=1):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER

def style_cell(cell, fill):
    cell.fill = fill; cell.border = BORDER
    cell.alignment = Alignment(vertical="center", wrap_text=True)

def ks_fill(ks):
    color = "D6EAD0" if ks == "kernel" else "FEF0E7"
    return PatternFill("solid", fgColor=color)

def ks_badge(ks):
    return "● KERNEL" if ks == "kernel" else "● SUPPORT"

# ─── Build workbook ───────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ══ TAB 1 – LIBRARIES ═════════════════════════════════════════════════════════
ws_lib = wb.active
ws_lib.title = "Libraries"
ws_lib.freeze_panes = "A2"

LIB_HEADERS = ["Filename (.cpp)", "Directory", "Group", "Sub-group",
               "Layer", "Kernel / Support", "Key COSMOS Dependencies",
               "Type (user)", "Notes / Issues"]
set_header_row(ws_lib, LIB_HEADERS)
ws_lib.row_dimensions[1].height = 30

for i, (name, directory, group, subgroup, layer, ks, deps, flag) in enumerate(LIBRARIES, start=2):
    row_fill = WARN_FILL if flag == "open" else layer_fill(layer)
    values = [name + ".cpp", directory, group, subgroup, layer,
              ks_badge(ks), deps, "",
              "⚠ Open issue" if flag == "open" else ""]
    for col, val in enumerate(values, 1):
        c = ws_lib.cell(row=i, column=col, value=val)
        style_cell(c, row_fill)
        if col == 5:
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")
        if col == 6:
            c.font = Font(bold=True, color="375623" if ks == "kernel" else "833C00")
            c.alignment = Alignment(horizontal="center", vertical="center")

for col, width in enumerate([28, 42, 10, 14, 8, 14, 58, 18, 22], 1):
    ws_lib.column_dimensions[get_column_letter(col)].width = width

legend_row = len(LIBRARIES) + 4
ws_lib.cell(row=legend_row, column=1, value="LEGEND").font = Font(bold=True)
for j, (color, text) in enumerate([
    ("D6EAD0", "K0–K5  KERNEL  – no dependency on get_cosmosresources(); runs without COSMOS resource tables"),
    ("FEF0E7", "S0–S4  SUPPORT – requires COSMOS resource files, disk I/O, or physics-simulation context"),
    ("FFE699", "⚠ Open issue – known dependency that may be stale or warrants further cleanup"),
]):
    r = legend_row + 1 + j
    c1 = ws_lib.cell(row=r, column=1, value="")
    c1.fill = PatternFill("solid", fgColor=color); c1.border = BORDER
    c2 = ws_lib.cell(row=r, column=2, value=text)
    c2.border = BORDER; c2.alignment = Alignment(vertical="center")

# ══ TAB 2 – PROGRAMS ══════════════════════════════════════════════════════════
ws_prg = wb.create_sheet("Programs")
ws_prg.freeze_panes = "A2"

PRG_HEADERS = ["Filename (.cpp)", "Directory", "Category", "Sub-Category",
               "Key COSMOS Dependencies", "Kernel / Support", "Type (user)", "Notes"]
set_header_row(ws_prg, PRG_HEADERS)
ws_prg.row_dimensions[1].height = 30

for i, (name, directory, ks, deps) in enumerate(PROGRAMS, start=2):
    fname = name + ".cpp"
    user_cat, user_sub = user_program_data.get(fname, (None, None))
    values = [fname, directory, user_cat or "", user_sub or "",
              deps, ks_badge(ks), "", ""]
    for col, val in enumerate(values, 1):
        c = ws_prg.cell(row=i, column=col, value=val)
        style_cell(c, ks_fill(ks))
        if col == 6:
            c.font = Font(bold=True, color="375623" if ks == "kernel" else "833C00")
            c.alignment = Alignment(horizontal="center", vertical="center")

for col, width in enumerate([30, 52, 12, 14, 55, 14, 18, 20], 1):
    ws_prg.column_dimensions[get_column_letter(col)].width = width

legend_row = len(PROGRAMS) + 4
ws_prg.cell(row=legend_row, column=1, value="LEGEND").font = Font(bold=True)
for j, (color, text) in enumerate([
    ("D6EAD0", "● KERNEL  – runs without COSMOS resource tables"),
    ("FEF0E7", "● SUPPORT – requires COSMOS resource tables (IERS, WMM, JPL ephemeris, etc.)"),
], start=1):
    r = legend_row + j
    c1 = ws_prg.cell(row=r, column=1, value="")
    c1.fill = PatternFill("solid", fgColor=color); c1.border = BORDER
    c2 = ws_prg.cell(row=r, column=2, value=text)
    c2.border = BORDER; c2.alignment = Alignment(vertical="center")

# ─── Save ──────────────────────────────────────────────────────────────────────
out = os.path.join(BASE, "cosmos_cpp_inventory.xlsx")
wb.save(out)
print(f"Saved: {out}")
print(f"  Libraries tab: {len(LIBRARIES)} entries")
print(f"  Programs tab:  {len(PROGRAMS)} entries")
print(f"  User program data preserved: {len(user_program_data)} rows")
